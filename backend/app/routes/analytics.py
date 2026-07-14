from fastapi import APIRouter, Depends, Query
from datetime import datetime, date, timedelta
from typing import Optional, List
from app.core.database import get_db
from app.core.security import get_current_user, require_admin_or_manager
from app.models import TaskUpload, TaskSession, User
from app.services.utils import seconds_to_hours

router = APIRouter(prefix="/api/analytics", tags=["Analytics"])


def get_leave_days_in_period(user_id: int, start: date, end: date, db) -> int:
    """
    Count weekdays of leave for a user within [start, end].
    All submitted leaves automatically adjust productivity — no approval step needed.
    """
    query = {
        "user_id": user_id,
        "from_date": {"$lte": str(end)},
        "to_date": {"$gte": str(start)},
    }
    leaves_cursor = db.leaves.find(query)
    total_days = 0
    for leave in leaves_cursor:
        from_d = date.fromisoformat(leave["from_date"]) if isinstance(leave["from_date"], str) else leave["from_date"]
        to_d = date.fromisoformat(leave["to_date"]) if isinstance(leave["to_date"], str) else leave["to_date"]
        l_start = max(start, from_d)
        l_end = min(end, to_d)
        curr = l_start
        while curr <= l_end:
            if curr.weekday() < 5:
                total_days += 1
            curr += timedelta(days=1)
    return total_days
@router.get("/dashboard/me")
def my_dashboard(
    db = Depends(get_db),
    current_user=Depends(get_current_user),
):
    today = date.today()
    week_start = today - timedelta(days=today.weekday())

    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())

    # Today's tasks
    today_tasks_cursor = db.task_uploads.find({
        "user_id": current_user.id,
        "created_at": {"$gte": today_start, "$lte": today_end}
    })
    today_tasks = [TaskUpload(**t) for t in today_tasks_cursor]

    # This week total seconds
    user_task_ids = [t["id"] for t in db.task_uploads.find({"user_id": current_user.id}, {"id": 1})]
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_sessions_cursor = db.task_sessions.find({
        "task_id": {"$in": user_task_ids},
        "started_at": {"$gte": week_start_dt}
    })
    week_seconds = sum(s.get("duration_seconds") or 0 for s in week_sessions_cursor)

    # Active task
    active_task_dict = db.task_uploads.find_one({
        "user_id": current_user.id,
        "timer_status": "active",
    })
    active_task = TaskUpload(**active_task_dict) if active_task_dict else None

    # Active task current live seconds
    active_seconds = 0
    if active_task:
        open_session = db.task_sessions.find_one({
            "task_id": active_task.id,
            "ended_at": None,
            "paused_at": None,
        })
        if open_session:
            started_at = open_session["started_at"]
            if isinstance(started_at, str):
                started_at = datetime.fromisoformat(started_at)
            active_seconds = int((datetime.utcnow() - started_at).total_seconds())

    # WIP and paused tasks
    wip_cursor = db.task_uploads.find({
        "user_id": current_user.id,
        "timer_status": {"$in": ["active", "paused"]},
    })
    wip_tasks = [TaskUpload(**t) for t in wip_cursor]

    # Today's tasks (created today)
    today_tasks_cursor = db.task_uploads.find({
        "user_id": current_user.id,
        "created_at": {"$gte": today_start, "$lte": today_end}
    })
    today_tasks = [TaskUpload(**t) for t in today_tasks_cursor]

    # Today's time breakdown per task
    today_sessions_cursor = db.task_sessions.find({
        "task_id": {"$in": user_task_ids},
        "started_at": {"$gte": today_start, "$lte": today_end}
    })
    today_sessions = [TaskSession(**s) for s in today_sessions_cursor]
    
    sessions_by_task = {}
    for s in today_sessions:
        sessions_by_task[s.task_id] = sessions_by_task.get(s.task_id, 0) + (s.duration_seconds or 0)

    all_today_task_ids = set(sessions_by_task.keys())
    if active_task:
        all_today_task_ids.add(active_task.id)
    for t in today_tasks:
        if t.total_seconds > 0:
            all_today_task_ids.add(t.id)

    today_breakdown = []
    for tid in all_today_task_ids:
        task_dict = db.task_uploads.find_one({"id": tid})
        if task_dict:
            task = TaskUpload(**task_dict)
            today_secs = sessions_by_task.get(tid, 0)
            
            # If no sessions exist today, but the task was created today, sum its total_seconds
            if today_secs == 0 and task.id in [tk.id for tk in today_tasks]:
                today_secs = task.total_seconds

            if task.timer_status == "active" and active_task and task.id == active_task.id:
                today_secs += active_seconds
            if today_secs > 0:
                today_breakdown.append({
                    "task_id": task.id,
                    "ticket_id": task.ticket_id,
                    "task_title": task.task_title,
                    "track": task.track,
                    "seconds_today": today_secs,
                    "hours_today": seconds_to_hours(today_secs),
                    "status": task.timer_status,
                })

    completed_today = db.task_uploads.count_documents({
        "user_id": current_user.id,
        "status": "completed",
        "completed_at": {"$gte": today_start, "$lte": today_end}
    })

    return {
        "user": current_user.username,
        "today_date": str(today),
        "kpis": {
            "total_seconds_today": sum(t["seconds_today"] for t in today_breakdown),
            "total_hours_today": seconds_to_hours(sum(t["seconds_today"] for t in today_breakdown)),
            "total_hours_this_week": seconds_to_hours(week_seconds),
            "wip_count": len(wip_tasks),
            "active_count": sum(1 for t in wip_tasks if t.timer_status == "active"),
            "paused_count": sum(1 for t in wip_tasks if t.timer_status == "paused"),
            "completed_today": completed_today,
        },
        "active_task": {
            "task_id": active_task.id,
            "ticket_id": active_task.ticket_id,
            "task_title": active_task.task_title,
            "track": active_task.track,
            "total_seconds": active_task.total_seconds,
            "current_session_seconds": active_seconds,
        } if active_task else None,
        "today_breakdown": sorted(today_breakdown, key=lambda x: x["seconds_today"], reverse=True),
        "wip_tasks": [
            {
                "task_id": t.id,
                "ticket_id": t.ticket_id,
                "task_title": t.task_title,
                "track": t.track,
                "priority": t.priority,
                "timer_status": t.timer_status,
                "total_seconds": t.total_seconds,
                "total_hours": seconds_to_hours(t.total_seconds),
            }
            for t in wip_tasks
        ],
    }


@router.get("/dashboard/admin")
def admin_dashboard(
    view_type: str = "weekly",
    week_offset: int = 0,
    offset: Optional[int] = None,
    db = Depends(get_db),
    _=Depends(require_admin_or_manager),
):
    actual_offset = offset if offset is not None else week_offset
    today = date.today()

    if view_type == "monthly":
        year = today.year
        month = today.month - actual_offset
        while month <= 0:
            month += 12
            year -= 1
        
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        label = start_date.strftime("%B %Y")
    else:
        start_date = today - timedelta(days=today.weekday()) - timedelta(weeks=actual_offset)
        end_date = start_date + timedelta(days=6)
        label = f"{start_date} to {end_date}"

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    # 1. Get all developers
    developers_cursor = db.users.find({
        "role": {"$in": ["developer", "intern"]},
        "is_active": True,
    })
    developers = [User(**u) for u in developers_cursor]

    # Gather tasks and sessions across all developers
    tasks_in_period = []
    dev_breakdown = []
    total_seconds = 0
    status_counts = {}
    track_counts = {}

    for dev in developers:
        # Get all sessions in this period for this developer
        sessions_cursor = db.task_sessions.find({
            "user_id": dev.id,
            "started_at": {"$gte": start_dt, "$lte": end_dt},
        })
        sessions = [TaskSession(**s) for s in sessions_cursor]
        task_ids_with_sessions = set(s.task_id for s in sessions)

        # Get all manual tasks for this developer in this period
        dev_tasks_cursor = db.task_uploads.find({
            "user_id": dev.id,
            "start_date": {"$gte": str(start_date), "$lte": str(end_date)}
        })
        dev_tasks = [TaskUpload(**t) for t in dev_tasks_cursor]

        # Load all tasks that had sessions this week
        session_task_ids = list(task_ids_with_sessions)
        session_tasks = []
        if session_task_ids:
            session_tasks = [TaskUpload(**t) for t in db.task_uploads.find({"id": {"$in": session_task_ids}})]

        # Unique tasks for this developer this week
        all_dev_week_tasks = {t.id: t for t in dev_tasks + session_tasks}.values()
        tasks_in_period.extend(all_dev_week_tasks)

        dev_total_secs = 0
        for t in all_dev_week_tasks:
            # 1. Sum of sessions for this task in the current week
            task_week_sessions = [s for s in sessions if s.task_id == t.id]
            task_week_secs = sum(s.duration_seconds or 0 for s in task_week_sessions)

            # 2. Calculate manual seconds
            all_sessions_cursor = db.task_sessions.find({"task_id": t.id})
            all_sessions_secs = sum(s.get("duration_seconds") or 0 for s in all_sessions_cursor)
            
            manual_secs = max(0, (t.total_seconds or 0) - all_sessions_secs)

            # If task's start_date is in this week, count manual seconds
            is_start_date_in_week = False
            if t.start_date:
                t_start = t.start_date
                if isinstance(t_start, str):
                    try:
                        t_start = date.fromisoformat(t_start)
                    except:
                        pass
                if isinstance(t_start, date) and start_date <= t_start <= end_date:
                    is_start_date_in_week = True

            task_total_secs = task_week_secs
            if is_start_date_in_week:
                task_total_secs += manual_secs

            dev_total_secs += task_total_secs

        total_seconds += dev_total_secs

        dev_breakdown.append({
            "user_id": dev.id,
            "username": dev.username,
            "full_name": dev.full_name,
            "role": dev.role,
            "dev_type": dev.dev_type,
            "total_seconds": dev_total_secs,
            "total_hours": seconds_to_hours(dev_total_secs),
            "task_count": len(all_dev_week_tasks),
            "leave_days": get_leave_days_in_period(dev.id, start_date, end_date, db),
        })

    # Sort developer breakdown by total seconds logged descending
    dev_breakdown.sort(key=lambda x: x["total_seconds"], reverse=True)

    # Unique tasks overall
    unique_tasks = {t.id: t for t in tasks_in_period}.values()

    # Calculate status and track counts from the list of active tasks this week
    for task in unique_tasks:
        status_counts[task.status] = status_counts.get(task.status, 0) + 1
        if task.track:
            track_counts[task.track] = track_counts.get(task.track, 0) + 1

    total_users = db.users.count_documents({"is_active": True})

    return {
        "week_label": label,
        "kpis": {
            "total_users": total_users,
            "total_tasks_this_week": len(unique_tasks),
            "completed_this_week": status_counts.get("completed", 0),
            "in_progress_this_week": status_counts.get("in_progress", 0),
            "total_hours_this_week": seconds_to_hours(total_seconds),
        },
        "status_breakdown": status_counts,
        "track_breakdown": track_counts,
        "developer_breakdown": dev_breakdown,
    }


@router.get("/reports")
def export_report_data(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    track: Optional[str] = None,
    db = Depends(get_db),
    current_user=Depends(get_current_user),
):
    filt = {}
    if current_user.role in ("developer", "intern"):
        filt["user_id"] = current_user.id
    elif user_id:
        filt["user_id"] = user_id

    if from_date or to_date:
        date_query = {}
        if from_date:
            date_query["$gte"] = str(from_date)
        if to_date:
            date_query["$lte"] = str(to_date)
        filt["start_date"] = date_query

    if status:
        filt["status"] = status
    if track:
        filt["track"] = track

    cursor = db.task_uploads.find(filt).sort("created_at", -1)
    tasks = [TaskUpload(**t) for t in cursor]

    total_seconds = sum(t.total_seconds or 0 for t in tasks)

    return {
        "total_tasks": len(tasks),
        "total_hours": seconds_to_hours(total_seconds),
        "avg_hours_per_task": seconds_to_hours(total_seconds // len(tasks)) if tasks else 0,
        "tasks": [
            {
                "ticket_id": t.ticket_id,
                "developer_name": t.developer_name,
                "task_title": t.task_title,
                "track": t.track,
                "dev_type": t.dev_type_task,
                "type_of_development": t.type_of_development,
                "cd_number": t.cd_number,
                "functional_team": t.functional_team,
                "status": t.status,
                "priority": t.priority,
                "start_date": str(t.start_date) if t.start_date else None,
                "due_date": str(t.due_date) if t.due_date else None,
                "hours_logged": float(t.hours_logged or 0),
                "total_seconds": t.total_seconds,
                "upload_source": t.upload_source,
                "created_at": str(t.created_at),
                "module": t.module,
                "category": t.category,
                "remarks": t.remarks,
                "description": t.description,
            }
            for t in tasks
        ],
    }


@router.get("/weekly-productivity")
def weekly_productivity(
    offset: int = 0,
    db = Depends(get_db),
    current_user=Depends(get_current_user),
):
    today = date.today()
    start_date = today - timedelta(days=today.weekday()) - timedelta(weeks=offset)
    end_date = start_date + timedelta(days=6)
    
    day = start_date.day
    ordinal = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}.get((day - 1) // 7 + 1, "1st")
    month_name = start_date.strftime("%B")
    week_label = f"{ordinal} week of {month_name} {start_date.year}"

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    developers_cursor = db.users.find({
        "role": {"$in": ["developer", "intern"]},
        "is_active": True,
    })
    developers = [User(**u) for u in developers_cursor]

    breakdown = []
    for dev in developers:
        # 1. Get all task sessions in this week for this developer
        sessions_cursor = db.task_sessions.find({
            "user_id": dev.id,
            "started_at": {"$gte": start_dt, "$lte": end_dt},
        })
        sessions = [TaskSession(**s) for s in sessions_cursor]
        task_ids_with_sessions = set(s.task_id for s in sessions)

        # 2. Get all manual tasks for this developer with start_date within this week
        dev_tasks_cursor = db.task_uploads.find({
            "user_id": dev.id,
            "start_date": {"$gte": str(start_date), "$lte": str(end_date)}
        })
        dev_tasks = [TaskUpload(**t) for t in dev_tasks_cursor]

        # Get all tasks that had sessions this week
        session_task_ids = list(task_ids_with_sessions)
        session_tasks = []
        if session_task_ids:
            session_tasks = [TaskUpload(**t) for t in db.task_uploads.find({"id": {"$in": session_task_ids}})]
            
        # Combine the lists and remove duplicates
        all_week_tasks = {t.id: t for t in dev_tasks + session_tasks}.values()

        total_secs = 0
        weekend_secs = 0
        task_list = []
        for t in all_week_tasks:
            # 1. Sum sessions for this task in the period — split by weekday vs weekend
            task_week_sessions = [s for s in sessions if s.task_id == t.id]
            task_weekday_secs = 0
            task_weekend_secs = 0
            for s in task_week_sessions:
                started = s.started_at
                if isinstance(started, str):
                    try:
                        started = datetime.fromisoformat(started)
                    except:
                        pass
                if isinstance(started, datetime) and started.weekday() >= 5:  # Sat=5, Sun=6
                    task_weekend_secs += (s.duration_seconds or 0)
                else:
                    task_weekday_secs += (s.duration_seconds or 0)
            task_week_secs = task_weekday_secs + task_weekend_secs

            # 2. Calculate manual seconds: total_seconds minus all session seconds across all time
            all_sessions_cursor = db.task_sessions.find({"task_id": t.id})
            all_sessions_secs = sum(s.get("duration_seconds") or 0 for s in all_sessions_cursor)
            
            manual_secs = max(0, (t.total_seconds or 0) - all_sessions_secs)

            # If task's start_date is in this week, count manual seconds
            is_start_date_in_week = False
            if t.start_date:
                t_start = t.start_date
                if isinstance(t_start, str):
                    try:
                        t_start = date.fromisoformat(t_start)
                    except:
                        pass
                if isinstance(t_start, date) and start_date <= t_start <= end_date:
                    is_start_date_in_week = True

            task_total_secs = task_week_secs
            if is_start_date_in_week:
                task_total_secs += manual_secs

            total_secs += task_total_secs
            weekend_secs += task_weekend_secs
            task_mins = int(round(task_total_secs / 60))

            # Detect if the task has any weekend sessions
            is_weekend_task = task_weekend_secs > 0

            task_list.append({
                "id": t.id,
                "ticket_id": t.ticket_id,
                "task_title": t.task_title,
                "track": t.track,
                "status": t.status,
                "hours_logged": round(task_mins / 60, 2),
                "weekend_hours": round(task_weekend_secs / 3600, 2),
                "is_weekend_work": is_weekend_task,
                "created_at": str(t.created_at),
            })

        total_mins = int(round(total_secs / 60))
        weekend_mins = int(round(weekend_secs / 60))
        hrs = total_mins // 60
        mins = total_mins % 60
        hours_str = f"{hrs} hours {mins} minutes"

        w_hrs = weekend_mins // 60
        w_mins = weekend_mins % 60
        weekend_hours_str = f"{w_hrs} hours {w_mins} minutes" if weekend_mins > 0 else ""

        # All leave days automatically reduce the 40-hr (2400 min) weekly target
        leave_days = get_leave_days_in_period(dev.id, start_date, end_date, db)
        target_mins = max(2400 - (leave_days * 480), 0)  # 480 min = 8 hrs/day
        target_hrs = round(target_mins / 60, 1)

        if target_mins > 0:
            prod_pct = int(round((total_mins / target_mins) * 100))
        else:
            prod_pct = 100  # Developer on full-week leave

        breakdown.append({
            "developer_id": dev.id,
            "username": dev.username,
            "full_name": dev.full_name or dev.username,
            "dev_type": dev.dev_type,
            "total_minutes": total_mins,
            "hours_str": hours_str,
            "weekend_minutes": weekend_mins,
            "weekend_hours_str": weekend_hours_str,
            "productivity_pct": prod_pct,
            "tasks": task_list,
            "leave_days": leave_days,
            "target_hours": target_hrs,
            "target_minutes": target_mins,
        })

    breakdown.sort(key=lambda x: x["full_name"].lower())

    return {
        "week_label": week_label,
        "date_range": f"{start_date} to {end_date}",
        "data": breakdown,
    }


def _compute_productivity_for_range(start_date: date, end_date: date, label: str, db) -> dict:
    """Shared helper: compute productivity breakdown for all active developers in [start_date, end_date]."""
    today = date.today()
    # Cap end_date to today so we don't report future days
    effective_end = min(end_date, today)

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(effective_end, datetime.max.time())

    # Count weekdays in the period for target calculation
    def count_weekdays(s: date, e: date) -> int:
        total = 0
        cur = s
        while cur <= e:
            if cur.weekday() < 5:
                total += 1
            cur += timedelta(days=1)
        return total

    period_weekdays = count_weekdays(start_date, effective_end)
    period_target_mins_base = period_weekdays * 480  # 8 hrs per weekday

    developers_cursor = db.users.find({
        "role": {"$in": ["developer", "intern"]},
        "is_active": True,
    })
    developers = [User(**u) for u in developers_cursor]

    breakdown = []
    for dev in developers:
        sessions_cursor = db.task_sessions.find({
            "user_id": dev.id,
            "started_at": {"$gte": start_dt, "$lte": end_dt},
        })
        sessions = [TaskSession(**s) for s in sessions_cursor]
        task_ids_with_sessions = set(s.task_id for s in sessions)

        dev_tasks_cursor = db.task_uploads.find({
            "user_id": dev.id,
            "start_date": {"$gte": str(start_date), "$lte": str(effective_end)}
        })
        dev_tasks = [TaskUpload(**t) for t in dev_tasks_cursor]

        session_task_ids = list(task_ids_with_sessions)
        session_tasks = []
        if session_task_ids:
            session_tasks = [TaskUpload(**t) for t in db.task_uploads.find({"id": {"$in": session_task_ids}})]

        all_period_tasks = {t.id: t for t in dev_tasks + session_tasks}.values()

        total_secs = 0
        task_list = []
        for t in all_period_tasks:
            task_period_sessions = [s for s in sessions if s.task_id == t.id]
            task_period_secs = sum(s.duration_seconds or 0 for s in task_period_sessions)

            all_sessions_cursor = db.task_sessions.find({"task_id": t.id})
            all_sessions_secs = sum(s.get("duration_seconds") or 0 for s in all_sessions_cursor)
            manual_secs = max(0, (t.total_seconds or 0) - all_sessions_secs)

            is_start_date_in_period = False
            if t.start_date:
                t_start = t.start_date
                if isinstance(t_start, str):
                    try:
                        t_start = date.fromisoformat(t_start)
                    except:
                        pass
                if isinstance(t_start, date) and start_date <= t_start <= effective_end:
                    is_start_date_in_period = True

            task_total_secs = task_period_secs
            if is_start_date_in_period:
                task_total_secs += manual_secs

            total_secs += task_total_secs
            task_mins = int(round(task_total_secs / 60))

            task_list.append({
                "id": t.id,
                "ticket_id": t.ticket_id,
                "task_title": t.task_title,
                "track": t.track,
                "status": t.status,
                "hours_logged": round(task_mins / 60, 2),
                "created_at": str(t.created_at),
            })

        total_mins = int(round(total_secs / 60))
        hrs = total_mins // 60
        mins = total_mins % 60
        hours_str = f"{hrs} hours {mins} minutes"

        leave_days = get_leave_days_in_period(dev.id, start_date, effective_end, db)
        target_mins = max(period_target_mins_base - (leave_days * 480), 0)
        target_hrs = round(target_mins / 60, 1)

        if target_mins > 0:
            prod_pct = int(round((total_mins / target_mins) * 100))
        else:
            prod_pct = 100

        breakdown.append({
            "developer_id": dev.id,
            "username": dev.username,
            "full_name": dev.full_name or dev.username,
            "dev_type": dev.dev_type,
            "total_minutes": total_mins,
            "hours_str": hours_str,
            "productivity_pct": prod_pct,
            "tasks": task_list,
            "leave_days": leave_days,
            "target_hours": target_hrs,
            "target_minutes": target_mins,
        })

    breakdown.sort(key=lambda x: x["full_name"].lower())

    return {
        "period_label": label,
        "date_range": f"{start_date} to {effective_end}",
        "period_weekdays": period_weekdays,
        "data": breakdown,
    }


@router.get("/monthly-productivity")
def monthly_productivity(
    offset: int = 0,
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Return productivity for a calendar month, capped at today for the current month."""
    today = date.today()
    year = today.year
    month = today.month - offset
    while month <= 0:
        month += 12
        year -= 1

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    label = start_date.strftime("%B %Y")
    return _compute_productivity_for_range(start_date, end_date, label, db)


@router.get("/custom-productivity")
def custom_productivity(
    from_date: date = Query(...),
    to_date: date = Query(...),
    db=Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Return productivity for an arbitrary date range."""
    if to_date < from_date:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="to_date cannot be before from_date")
    label = f"{from_date.strftime('%d %b %Y')} – {to_date.strftime('%d %b %Y')}"
    return _compute_productivity_for_range(from_date, to_date, label, db)


@router.get("/weekly-productivity/export")
def export_weekly_productivity(
    offset: int = 0,
    db = Depends(get_db),
    _=Depends(require_admin_or_manager),
):
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    start_date = today - timedelta(days=today.weekday()) - timedelta(weeks=offset)
    end_date = start_date + timedelta(days=6)
    
    day = start_date.day
    ordinal = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}.get((day - 1) // 7 + 1, "1st")
    month_name = start_date.strftime("%B")
    week_label = f"{ordinal} week of {month_name} {start_date.year}"

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    developers_cursor = db.users.find({
        "role": {"$in": ["developer", "intern"]},
        "is_active": True,
    })
    developers = [User(**u) for u in developers_cursor]

    wb = Workbook()
    ws = wb.active
    ws.title = "Productivity Summary"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=11, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True)
    font_data = Font(name="Calibri", size=11)
    font_total = Font(name="Calibri", size=11, bold=True)

    fill_title = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    border_total = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="double", color="000000")
    )

    ws.merge_cells("A1:E1")
    ws["A1"] = week_label
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = fill_title

    ws.merge_cells("A2:E2")
    ws["A2"] = "Developers Weekly (40 hours) productivity"
    ws["A2"].font = font_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = fill_title

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    headers = ["Row Labels", "Sum of Time (Min)", "Sum of Time (Hours)", "Productivity", "Leaves (Days)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left" if col_num in (1, 3) else "right", vertical="center")
        cell.border = border_data

    ws.row_dimensions[3].height = 20

    row_idx = 4
    for dev in sorted(developers, key=lambda x: x.full_name or x.username):
        # 1. Get all task sessions in this week for this developer
        sessions_cursor = db.task_sessions.find({
            "user_id": dev.id,
            "started_at": {"$gte": start_dt, "$lte": end_dt},
        })
        sessions = [TaskSession(**s) for s in sessions_cursor]
        task_ids_with_sessions = set(s.task_id for s in sessions)

        # 2. Get all manual tasks for this developer with start_date within this week
        dev_tasks_cursor = db.task_uploads.find({
            "user_id": dev.id,
            "start_date": {"$gte": str(start_date), "$lte": str(end_date)}
        })
        dev_tasks = [TaskUpload(**t) for t in dev_tasks_cursor]

        # Get all tasks that had sessions this week
        session_task_ids = list(task_ids_with_sessions)
        session_tasks = []
        if session_task_ids:
            session_tasks = [TaskUpload(**t) for t in db.task_uploads.find({"id": {"$in": session_task_ids}})]
            
        # Combine the lists and remove duplicates
        all_week_tasks = {t.id: t for t in dev_tasks + session_tasks}.values()

        total_secs = 0
        for t in all_week_tasks:
            # 1. Sum of sessions for this task in the current week
            task_week_sessions = [s for s in sessions if s.task_id == t.id]
            task_week_secs = sum(s.duration_seconds or 0 for s in task_week_sessions)

            # 2. Calculate manual seconds
            all_sessions_cursor = db.task_sessions.find({"task_id": t.id})
            all_sessions_secs = sum(s.get("duration_seconds") or 0 for s in all_sessions_cursor)
            
            manual_secs = max(0, (t.total_seconds or 0) - all_sessions_secs)

            # If task's start_date is in this week, count manual seconds
            is_start_date_in_week = False
            if t.start_date:
                t_start = t.start_date
                if isinstance(t_start, str):
                    try:
                        t_start = date.fromisoformat(t_start)
                    except:
                        pass
                if isinstance(t_start, date) and start_date <= t_start <= end_date:
                    is_start_date_in_week = True

            task_total_secs = task_week_secs
            if is_start_date_in_week:
                task_total_secs += manual_secs

            total_secs += task_total_secs

        total_mins = int(round(total_secs / 60))

        ws.cell(row=row_idx, column=1, value=dev.full_name or dev.username).font = font_data
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=1).border = border_data

        ws.cell(row=row_idx, column=2, value=total_mins).font = font_data
        ws.cell(row=row_idx, column=2).number_format = "#,##0"
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=2).border = border_data

        hrs = total_mins // 60
        mins = total_mins % 60
        hours_str = f"{hrs} hours {mins} minutes"
        ws.cell(row=row_idx, column=3, value=hours_str).font = font_data
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=3).border = border_data

        # All leave days automatically reduce the 40-hr target
        leave_days_export = get_leave_days_in_period(dev.id, start_date, end_date, db)
        target_mins = max(2400 - (leave_days_export * 480), 1)  # avoid division by zero
        prod_pct_val = round(total_mins / target_mins, 4)
        ws.cell(row=row_idx, column=4, value=prod_pct_val).font = font_data
        ws.cell(row=row_idx, column=4).number_format = "0%"
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=4).border = border_data

        ws.cell(row=row_idx, column=5, value=leave_days_export).font = font_data
        ws.cell(row=row_idx, column=5).number_format = "#,##0"
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=5).border = border_data

        row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value="Grand Total").font = font_total
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=total_row, column=1).border = border_total

    ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = font_total
    ws.cell(row=total_row, column=2).number_format = "#,##0"
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).border = border_total

    ws.cell(row=total_row, column=3, value="").font = font_total
    ws.cell(row=total_row, column=3).border = border_total

    ws.cell(row=total_row, column=4, value=f"=B{total_row}/(COUNT(B4:B{total_row-1})*2400)").font = font_total
    ws.cell(row=total_row, column=4).number_format = "0%"
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).border = border_total

    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})").font = font_total
    ws.cell(row=total_row, column=5).number_format = "#,##0"
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=5).border = border_total

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Weekly_Productivity_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
